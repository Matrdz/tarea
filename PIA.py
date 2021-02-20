import os, requests, re
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook

urls=list()
varlist=list()
inf=dict()
noti=dict()
os.makedirs("Imagenes", exist_ok=True)
archivos=os.listdir()
for i in archivos:
    if ".txt" in i:
        with open (str(i), "r") as file:
            for line in file:
                if "https" in line:
                    #print("ulrs en",str(i))
                    urls.append(line)
#input("pausa")
##print(urls)
#input("pausa")
if "biografiasyvidas" in urls[0]:
    url=str(urls[0])
    print("Descargando pagina: "+url)
    url=url.rstrip("\n")
    page = requests.get (url)
    url=url.rstrip("messi.htm")
    if page.status_code != 200:
        print("Pagina no encontrada")
    else:
        soup=bs(page.text, "html.parser")
        po=soup.find_all("p", class_="piefotos")
        if po == []:
            print('No se encontró.')
        else:
            for x in (0,1,3):
                imgurl = po[x].contents[0].get("src")
                print("Descargando %s..." % (imgurl))
                response = requests.get(url+imgurl)
                if response.status_code == 200:
                    imageFile = open(os.path.join("Imagenes", os.path.basename(imgurl)), "wb")
                    for chunk in response.iter_content(100000):
                        imageFile.write(chunk)
                    imageFile.close()
        parra=soup.select("p")
        var=(parra[20].getText()) #buscando el equipo
        patron=re.compile(r"Messi en el F.C. ([a-zA-Z0-9]+)")
        mo=patron.search(var)
        inf["Futbol Club"]=mo.group(1)
        
#input("pausa")
if "lavanguardia" in urls[1]:
    url=str(urls[1])
    url=url.rstrip("\n")
    print("\n\nDescargando pagina: "+url)
    page = requests.get (url)
    if page.status_code != 200:
        print("Pagina no encontrada")
    else:
        soup=bs(page.content,"html.parser") #buscando las fechas
        po=soup.find_all ("p")
        text1=po[5].getText()
        #print(text1)
        text2=po[7].getText()
        #print(text2)
        patron=re.compile(r"(\d*\d) de diciembre \((\d\d):(\d\d) horas\)")
        mo1=patron.findall(text1)
        mo2=patron.findall(text2)
        patron=re.compile(r"(\d*\d) \((\d\d):(\d\d)( horas)*\)")
        mo3=patron.findall(text2)
        mo=mo1+mo2+mo3
        print(mo)
        file=open ("dias.txt", "w")
        for x in range(len(mo)):
            file.write("2020-12-")
            file.write(mo[x][0])
            file.write(" ")
            file.write(mo[x][1])
            file.write(":")
            file.write(mo[x][2])
            file.write("\n")
        file.close()
        print("fechas agregados correctamente")

if "guia-telefonica" in urls[2]:
    url=str(urls[2])
    print("\n\nDescargando pagina: "+url)
    url=url.rstrip("\n")
    page = requests.get (url)
    if page.status_code != 200:
        print("Pagina no encontrada")
    else:
        soup=bs(page.text,"html.parser") #buscando contactos
        po=soup.find_all ("p")
        i=0
        for x in range(len(po)):
            #print("\nnumero:",x)
            #print(po[x].getText())
            #print("\n\n")
            var=po[x].getText()
            patron=re.compile(r"fue fundado en el año ([0-9]+)")
            mo=patron.search(var)
            if mo != None:
                inf["Fundación"]=mo.group(1)
                print("Fundación encontrada!")
            patron=re.compile(r"Su estadio principal es el ([a-zA-Z0-9]+) ([a-zA-Z0-9]+)")
            mo=patron.search(var)
            if mo != None:
                var2=mo.group(1)+" "+mo.group(2)
                inf["Estadio Principal"]=var2
                print("Estadio encontrado!")
            patron=re.compile(r"www.([a-zA-Z0-9]+).([a-zA-Z0-9]+)")
            mo=patron.search(var)
            if mo != None:
                varlist.append(mo.group())
                print("Página encontrada!")
                inf["Paginas web"]=str(varlist)
        soup=bs(page.content,"html.parser")
        info=soup.select("p a")
        varlist=list()
        for etiqueta in info:
            url2=etiqueta.get("href")
            patron=re.compile(r"[a-zA-Z0-9]+@fcbarcelona.([a-zA-Z0-9]+)")
            mo=patron.search(url2)
            if mo != None:
                varlist.append(mo.group())
                print("Correo encontrado!")
                inf["Emails"]=str(varlist)
    libro=Workbook()
    hoja=libro.active
    libro.active=hoja
    hoja1=libro.create_sheet("Noti")
    hoja1=libro["Noti"]
    names= ("Dato", "-")
    i=1
    for x in names:
        hoja.cell(1,i, x)
        i+=1
    i=1
    for x in inf:
        i+=1
        #var=str(noti[x])
        hoja.cell(i,1, x) #x es la llave y inf[x] el valor
        hoja.cell(i,2, inf[x])
    libro.save("info.xlsx")
    print("Información guardada")

#input("pausa")
if "marca" in urls[3]: #noticias
    url=str(urls[3])
    print("Descargando pagina: "+url)
    url=url.rstrip("\n")
    page = requests.get (url)
    if page.status_code != 200:
        print("Pagina no encontrada")
    else:
        soup=bs(page.content,"html.parser")
        info=soup.select("h3 a")
        for etiqueta in info:
            title=etiqueta.get("title")
            url2=etiqueta.get("href")
            noti[title]=url2
        print("\n",len(noti),"Noticias encontradas!")
    libro=load_workbook("info.xlsx")
    hoja1=libro["Noti"]
    libro.active=hoja1
    hoja1.cell(1,1, "Noticias")
    hoja1.cell(1,10, "URL")
    i=1
    for x in noti:
        i+=1
        #var=str(noti[x])
        hoja1.cell(i,1, x)
        hoja1.cell(i,10, noti[x])
    libro.save("info.xlsx")
    print("\nNoticias guardadas")
#input("pausa")
print("Importando clima...")
import clima
if __name__=="__main__":
    print("\n\nclima guardado con éxito")








    

