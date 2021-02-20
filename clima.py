from datetime import datetime
import time, json, requests
from openpyxl import Workbook, load_workbook
dtemp=("day","night","max","min") #temperatura que me interesa
lat=("41.3592671","40.4381311","37.3754338") #ubicación de barcelona, madrid, sevilla
lon=("2.1599586","-3.8196193","-5.9900776")
days=(1,3,5)#mañana, en dos dias y en 4 dias
key=input("Inserta tu key")
now = str(datetime.now()) #la fecha
libro=load_workbook("info.xlsx")
hoja=libro["Sheet"]
hoja.cell(7,1,"Hoy")
hoja.cell(7,2,now)
hoja.cell(9,1,"Poximos juegos")
names=("dia","temp dia","temp noche","temp max","temp min","presión","húmedad")
i=1
for x in names:
    hoja.cell(10,i,x)
    i+=1

for x in range(3):
    url=("https://api.openweathermap.org/data/2.5/onecall?lat="+lat[x]+"&lon="+lon[x]+"&dt&exclude=minutely,hourly&units=metric&appid="+key)
    page = requests.get (url)
    wd = json.loads(page.content)
    dt=wd["daily"][days[x]]["dt"]#days[x] x=0 hoy x=1 mañana
    dt=datetime.utcfromtimestamp(dt).strftime("%d-%m-%Y")
    dt=str(dt)
    hoja.cell(11+x,1,dt)
    temp=wd["daily"][days[x]]["temp"]
    i=2
    for y in dtemp:
        var=(str(temp[y])+"°")
        hoja.cell(11+x,i,var)
        i+=1
        #print (y," = ",temp[y])
    press=wd["daily"][days[x]]["pressure"]
    var=(str(press)+"hPa")
    hoja.cell(11+x,6,var)
    
    hum=wd["daily"][days[x]]["humidity"]
    var=(str(hum)+"%")
    hoja.cell(11+x,7,var)
    #print("Clima guardado con éxito")
libro.save("info.xlsx")
